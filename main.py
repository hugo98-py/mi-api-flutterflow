# -*- coding: utf-8 -*-
"""
Created on Sat Apr 19 19:37:45 2025

@author: Hugo
"""
# from fastapi import FastAPI 

# app = FastAPI()

# # http://127.0.0.1:8000/

# @app.get("/")
# def index():
#     return "Hola, Pythonianos"

# Coduigo que funciona 
# from fastapi import FastAPI
# from fastapi.responses import StreamingResponse
# import pandas as pd
# from io import BytesIO

# app = FastAPI()

# @app.get("/")
# def index():
#     return {"mensaje": "Hola, Pythonianos"}

# @app.get("/crear-excel")
# def crear_excel(nombre: str = "Usuario"):
#     try:
#         # Crear DataFrame de ejemplo
#         datos = {
#             "Nombre": [nombre, nombre],
#             "Edad": [30, 40],
#             "Ciudad": ["Santiago", "Valparaíso"]
#         }
#         df = pd.DataFrame(datos)

#         # Guardar el Excel en memoria
#         output = BytesIO()
#         df.to_excel(output, index=False)
#         output.seek(0)

#         return StreamingResponse(
#             output,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": f"attachment; filename=datos_{nombre}.xlsx"}
#         )
    
#     except Exception as e:
#         return {"error": str(e)}



# Esta de aca tambien funciona. Permite modificar el excel planilla o molde y agrega 
# los datos input que entrego a la API
# from fastapi import FastAPI
# from fastapi.responses import StreamingResponse
# import pandas as pd
# from openpyxl import load_workbook
# from io import BytesIO

# app = FastAPI()

# @app.get("/modificar-excel")
# def modificar_excel(nombre: str = "Usuario"):
#     try:
#         # Cargar la plantilla existente
#         ruta_plantilla = "C:/Users/Hugo/OneDrive - AMS CONSULTORES SPA/Documentos/AMS/Proyectos/AMS App/fastapi/datos_Hugo.xlsx"
#         wb = load_workbook(ruta_plantilla)
#         ws = wb.active  # Usa la primera hoja (puedes especificar con wb["NombreHoja"])

#         # Agregar datos al final de la hoja
#         nueva_fila = [nombre, 30, "Santiago"]
#         ws.append(nueva_fila)

#         # Guardar en memoria
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)

#         return StreamingResponse(
#             output,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": f"attachment; filename=modificado_{nombre}.xlsx"}
#         )

#     except Exception as e:
#         return {"error": str(e)}


#ULTIMO 

# from fastapi import FastAPI
# from fastapi.responses import StreamingResponse
# from pydantic import BaseModel
# from openpyxl import load_workbook
# from io import BytesIO

# app = FastAPI()

# # ✅ Paso 1: Definir el modelo del JSON que recibirás
# class EntradaDatos(BaseModel):
#     nombre: str
#     edad: int
#     ciudad: str

# # ✅ Paso 2: Crear un endpoint POST para recibir el JSON
# @app.post("/modificar-excel")
# def modificar_excel(datos: EntradaDatos):
#     try:
#         # Ruta de tu archivo Excel de plantilla
#         ruta_plantilla = "C:/Users/Hugo/OneDrive - AMS CONSULTORES SPA/Documentos/AMS/Proyectos/AMS App/fastapi/datos_Hugo.xlsx"
        
#         # Cargar el archivo de plantilla
#         wb = load_workbook(ruta_plantilla)
#         ws = wb.active  # Puedes cambiarlo si tienes varias hojas

#         # Extraer los datos del JSON
#         nombre = datos.nombre
#         edad = datos.edad
#         ciudad = datos.ciudad

#         # Agregar una nueva fila al Excel con esos datos
#         nueva_fila = [nombre, edad, ciudad]
#         ws.append(nueva_fila)

#         # Guardar el archivo en memoria para enviarlo
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)

#         # Enviar el archivo como descarga
#         return StreamingResponse(
#             output,
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": f"attachment; filename=modificado_{nombre}.xlsx"}
#         )

#     except Exception as e:
#         return {"error": str(e)}


# Este fue el tultimo usado y el mas simple, unicamente para pruebas, funciona
# from fastapi import FastAPI
# from ngrok_asgi import NgrokMiddleware

# app = FastAPI()
# app.add_middleware(NgrokMiddleware)

# @app.get("/")
# def home():
#     return {"message": "Hola desde FastAPI + Ngrok"}

#from fastapi import FastAPI
#from pydantic import BaseModel

#app = FastAPI()

#class Datos(BaseModel):
#    nombre: str
#    edad: int

#@app.post("/crear")
#def crear_dato(data: Datos):
#    return {"mensaje": f"Hola {data.nombre}, tienes {data.edad} años."}

# Este es el codigo adaptado para generar un link de descarga con Post en FastAPI. (4 de mayo de 2025)
# pip install fastapi uvicorn openpyxl python-multipart
from fastapi import FastAPI
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import load_workbook
import uuid
import os

app = FastAPI()

# Cambiar al path de trabajo
#path = r'C:\Users\Hugo\OneDrive - AMS CONSULTORES SPA\Documentos\AMS\Proyectos\AMS App\fastapi'
#os.chdir(path)

# Asegurar que existe la carpeta static
os.makedirs("static", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Modelo para la solicitud
class DatosEntrada(BaseModel):
    nombre: str = "Sin nombre"
    edad: int | str = "N/A"

@app.post("/generar-excel/")
async def generar_excel(data: DatosEntrada):
    # Extraer datos
    nombre = data.nombre
    edad = data.edad

    # Cargar plantilla
    plantilla_path = "plantilla.xlsx"
    wb = load_workbook(plantilla_path)
    ws = wb.active

    # Escribir datos en la hoja
    ws["A1"] = "Nombre"
    ws["B1"] = "Edad"
    ws["A2"] = nombre
    ws["B2"] = edad

    # Generar archivo único
    archivo_id = str(uuid.uuid4())
    nombre_archivo = f"{archivo_id}.xlsx"
    ruta_guardado = os.path.join("static", nombre_archivo)
    wb.save(ruta_guardado)

    # URL de descarga
    url_descarga = f"http://localhost:8000/static/{nombre_archivo}"  # o cambia a tu dominio real en Render

    return JSONResponse({
        "message": "Archivo generado con éxito",
        "download_url": url_descarga
    })


